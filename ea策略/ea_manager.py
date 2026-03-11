# -*- coding: utf-8 -*-
"""
EA策略客户管理脚本
用法:
  python ea_manager.py list              # 列出所有客户
  python ea_manager.py list 进行中       # 按状态筛选
  python ea_manager.py query 焦伟林      # 查询某客户
  python ea_manager.py update 焦伟林 balance 12000u   # 更新字段
  python ea_manager.py add               # 新增客户（交互）
  python ea_manager.py export            # 导出 Excel
  python ea_manager.py summary           # 汇总统计
"""

import json
import sys
import os
from datetime import datetime

DATA_FILE = os.path.join(os.path.dirname(__file__), 'ea_clients.json')
EXCEL_FILE = os.path.join(os.path.dirname(__file__), 'EA策略资金表.xlsx')

def load():
    with open(DATA_FILE, 'r', encoding='utf-8') as f:
        return json.load(f)

def save(data):
    data['meta']['updated'] = datetime.now().strftime('%Y-%m-%d')
    # 重新计算合计
    active = sum(c['amount'] for c in data['clients'] if c['status'] == '进行中')
    stopped = sum(c['amount'] for c in data['clients'] if c['status'] == '停止')
    data['meta']['total_active'] = active
    data['meta']['total_stopped'] = stopped
    data['meta']['total'] = active + stopped
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    print(f"已保存，更新时间: {data['meta']['updated']}")

def cmd_list(args):
    data = load()
    status_filter = args[0] if args else None
    clients = data['clients']
    if status_filter:
        clients = [c for c in clients if c['status'] == status_filter]
    print(f"\n{'ID':<4} {'姓名':<16} {'负责人':<8} {'平台':<8} {'入金':<8} {'状态':<8} {'当前资金':<12} {'备注'}")
    print('-' * 80)
    for c in clients:
        platform = c.get('platform') or ''
        manager = c.get('manager') or ''
        balance = str(c.get('current_balance') or '')
        notes = c.get('notes') or ''
        print(f"{c['id']:<4} {c['name']:<16} {manager:<8} {platform:<8} "
              f"{c['amount']:<8} {c['status']:<8} {balance:<12} {notes}")
    print(f"\n共 {len(clients)} 条记录")

def cmd_query(args):
    if not args:
        print("用法: query <客户名>")
        return
    data = load()
    name = args[0]
    found = [c for c in data['clients'] if name in c['name']]
    if not found:
        print(f"未找到客户: {name}")
        return
    for c in found:
        print(f"\n{'='*40}")
        for k, v in c.items():
            print(f"  {k}: {v}")

def cmd_update(args):
    if len(args) < 3:
        print("用法: update <客户名> <字段> <新值>")
        print("可更新字段: status, current_balance, notes, platform, account, amount, transferred, manager")
        return
    data = load()
    name, field, value = args[0], args[1], args[2]
    found = [c for c in data['clients'] if name in c['name']]
    if not found:
        print(f"未找到客户: {name}")
        return
    # 类型转换
    if field == 'amount':
        value = int(value)
    elif field == 'transferred':
        value = value.lower() in ('true', '1', 'yes', '是')
    for c in found:
        old = c.get(field)
        c[field] = value
        print(f"已更新 [{c['name']}] {field}: {old} -> {value}")
    save(data)

def cmd_add(args):
    data = load()
    print("\n=== 新增客户 ===")
    new_id = max(c['id'] for c in data['clients']) + 1
    c = {
        'id': new_id,
        'name': input('姓名: '),
        'manager': input('负责人(张鹏/刘涛/曾涛/公司): '),
        'platform': input('平台(tmgm/dlsm/ebc): '),
        'account': input('账号ID: '),
        'amount': int(input('入金金额: ')),
        'status': input('状态(进行中/停止): '),
        'current_balance': input('当前资金: '),
        'transferred': input('是否已转(y/n): ').lower() == 'y',
        'notes': input('备注: ')
    }
    data['clients'].append(c)
    save(data)
    print(f"已新增客户: {c['name']} (ID: {new_id})")

def cmd_summary(args):
    data = load()
    print(f"\n{'='*40}")
    print(f"  更新时间: {data['meta']['updated']}")
    print(f"  进行中总计: {data['meta']['total_active']} U")
    print(f"  停止总计:   {data['meta']['total_stopped']} U")
    print(f"  总入金量:   {data['meta']['total']} U")
    print(f"\n  平台分布:")
    for p, v in data['platform_summary'].items():
        print(f"    {p}: {v}")
    print(f"\n  负责人分布:")
    managers = {}
    for c in data['clients']:
        m = c.get('manager', '未知')
        managers[m] = managers.get(m, 0) + c['amount']
    for m, v in sorted(managers.items(), key=lambda x: -x[1]):
        print(f"    {m}: {v} U")

def cmd_export(args):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    except ImportError:
        print("需要安装 openpyxl: pip install openpyxl")
        return

    data = load()
    wb = Workbook()
    ws = wb.active
    ws.title = "EA策略资金表"

    thin = Side(style='thin', color='000000')
    b = Border(left=thin, right=thin, top=thin, bottom=thin)

    def sc(row, col, value, bold=False, fill=None, size=11):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, size=size, name='Microsoft YaHei')
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = b
        if fill:
            c.fill = PatternFill(start_color=fill, end_color=fill, fill_type='solid')
        return c

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 18

    active_clients = [c for c in data['clients'] if c['status'] == '进行中']
    stopped_clients = [c for c in data['clients'] if c['status'] == '停止']

    row = 1
    ws.merge_cells(f'A{row}:C{row}')
    sc(row, 1, '进行中', bold=True, size=13, fill='E8F5E9')
    ws.row_dimensions[row].height = 26
    row += 1

    jx_start = row
    for c in active_clients:
        sc(row, 1, c.get('manager', ''))
        sc(row, 2, c['name'])
        sc(row, 3, c['amount'])
        ws.row_dimensions[row].height = 20
        row += 1
    jx_end = row - 1

    ws.merge_cells(start_row=jx_start, start_column=4, end_row=jx_end, end_column=4)
    cell = ws.cell(row=jx_start, column=4, value=f"共计：{data['meta']['total_active']}")
    cell.font = Font(bold=True, size=11, name='Microsoft YaHei')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = b

    ws.merge_cells(start_row=jx_start, start_column=5, end_row=jx_end + 1 + len(stopped_clients), end_column=5)
    cell = ws.cell(row=jx_start, column=5, value=f"总计入金量：{data['meta']['total']}U")
    cell.font = Font(bold=True, size=12, name='Microsoft YaHei')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = b

    ws.merge_cells(f'A{row}:C{row}')
    sc(row, 1, '停止', bold=True, size=13, fill='FFEBEE')
    ws.row_dimensions[row].height = 26
    row += 1

    tz_start = row
    for c in stopped_clients:
        sc(row, 1, c.get('manager', ''))
        sc(row, 2, c['name'])
        sc(row, 3, c['amount'])
        ws.row_dimensions[row].height = 20
        row += 1
    tz_end = row - 1

    ws.merge_cells(start_row=tz_start, start_column=4, end_row=tz_end, end_column=4)
    cell = ws.cell(row=tz_start, column=4, value=f"共计：{data['meta']['total_stopped']}")
    cell.font = Font(bold=True, size=11, name='Microsoft YaHei')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = b

    # 贵2
    row += 1
    ws.merge_cells(f'A{row}:C{row}')
    sc(row, 1, '贵2', bold=True, size=13, fill='FFF8E1')
    ws.row_dimensions[row].height = 26
    row += 1
    for g in data.get('gui2', []):
        sc(row, 1, '贵2')
        sc(row, 2, g['name'])
        sc(row, 3, g['amount'])
        ws.row_dimensions[row].height = 20
        row += 1

    wb.save(EXCEL_FILE)
    print(f"已导出: {EXCEL_FILE}")

if __name__ == '__main__':
    cmds = {'list': cmd_list, 'query': cmd_query, 'update': cmd_update,
            'add': cmd_add, 'summary': cmd_summary, 'export': cmd_export}
    if len(sys.argv) < 2 or sys.argv[1] not in cmds:
        print(__doc__)
        sys.exit(0)
    cmds[sys.argv[1]](sys.argv[2:])
