from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "EA策略资金表"

# 列宽
ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 18

# 边框样式
thin = Side(style='thin', color='000000')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def cell_style(ws, row, col, value, bold=False, align='center', fill_color=None, font_size=11):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=font_size, name='Microsoft YaHei')
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)
    c.border = border
    if fill_color:
        c.fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')
    return c

# ---- 第1行：表头 负责人 | 进行中（合并B-C）----
row = 1
cell_style(ws, row, 1, '负责人', bold=True)
ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=3)
cell_style(ws, row, 2, '进行中', bold=True, font_size=14)
ws.row_dimensions[row].height = 28

# ---- 进行中数据 ----
jinxing = [
    ('张鹏', '范宇轩', 2000),
    ('张鹏', '苏亮', 2000),
    ('张鹏', '雷熊', 2000),
    ('曾涛', '王欢', 2000),
    ('张鹏', '林锦豪', 4000),
    ('张鹏', '王泽远', 2000),
    ('张鹏', '梁杰煌', 2000),
    ('张鹏', '焦伟林', 10000),
    ('张鹏', '谢远然', 2000),
    ('公司', '雷蒙', 4000),
    ('张鹏', '聂平华', 6000),
    ('公司', '华哥（xiong yong）', 10000),
    ('刘涛', '孙成（chenyingying）', 2000),
    ('张鹏', '陈泽林', 2000),
    ('张鹏', '张金虎', 2000),
    ('刘涛', '杨静', 2000),
    ('张鹏', '杨浩', 2000),
]

start_row = 2
for i, (fzr, kh, je) in enumerate(jinxing):
    r = start_row + i
    cell_style(ws, r, 1, fzr, align='center')
    cell_style(ws, r, 2, kh, align='center')
    cell_style(ws, r, 3, je, align='center')
    ws.row_dimensions[r].height = 22

# 共计 58000 放在 D列，垂直居中合并
mid_row = start_row + len(jinxing) // 2
ws.merge_cells(start_row=start_row, start_column=4, end_row=start_row+len(jinxing)-1, end_column=4)
c = ws.cell(row=start_row, column=4, value='共计：58000')
c.font = Font(bold=True, size=11, name='Microsoft YaHei')
c.alignment = Alignment(horizontal='center', vertical='center')
c.border = border

# ---- 停止 标题行 ----
stop_title_row = start_row + len(jinxing)
ws.merge_cells(start_row=stop_title_row, start_column=1, end_row=stop_title_row, end_column=3)
cell_style(ws, stop_title_row, 1, '停止', bold=True, font_size=14)
ws.row_dimensions[stop_title_row].height = 28

# ---- 停止数据 ----
tingzhi = [
    ('刘涛', '雷伟平', 2000),
    ('刘涛', '朱睿', 2000),
    ('曾涛', '徐敏', 2000),
    ('张鹏', '孙明', 2000),
    ('张鹏', '李超杰', 2000),
    ('公司', '小叶', 8000),
    ('张鹏', '万诺', 2000),
    ('张鹏', '曾经，拥有', 2000),
    ('张鹏', '张凯嘉', 2000),
    ('张鹏', 'MN医疗', 2000),
    ('张鹏', 'bibiubiu', 2000),
    ('刘涛', '罗琴', 2000),
    ('张鹏', '徐宝龙', 2000),
]

tz_start = stop_title_row + 1
for i, (fzr, kh, je) in enumerate(tingzhi):
    r = tz_start + i
    cell_style(ws, r, 1, fzr, align='center')
    cell_style(ws, r, 2, kh, align='center')
    cell_style(ws, r, 3, je, align='center')
    ws.row_dimensions[r].height = 22

# 共计 32000
ws.merge_cells(start_row=tz_start, start_column=4, end_row=tz_start+len(tingzhi)-1, end_column=4)
c = ws.cell(row=tz_start, column=4, value='共计：32000')
c.font = Font(bold=True, size=11, name='Microsoft YaHei')
c.alignment = Alignment(horizontal='center', vertical='center')
c.border = border

# 总计入金量 在 E 列，对应进行中+停止中间位置
total_row = start_row + len(jinxing) // 2 + 2
ws.merge_cells(start_row=start_row, start_column=5, end_row=tz_start+len(tingzhi)-1, end_column=5)
c = ws.cell(row=start_row, column=5, value='总计入金量：90000U')
c.font = Font(bold=True, size=12, name='Microsoft YaHei')
c.alignment = Alignment(horizontal='center', vertical='center')
c.border = border

# ---- 空行分隔 ----
empty_row = tz_start + len(tingzhi)
ws.row_dimensions[empty_row].height = 10

# ---- 贵2数据 ----
gui2 = [
    ('贵2', '李维', 2000),
    ('贵2', '刘学习', 10000),
    ('贵2', '刘学习推荐', 4000),
]
gui2_start = empty_row + 1
for i, (fzr, kh, je) in enumerate(gui2):
    r = gui2_start + i
    c1 = cell_style(ws, r, 1, fzr, align='center')
    if i == 0:
        c1.border = Border(left=Side(style='medium', color='0070C0'),
                           right=thin, top=Side(style='medium', color='0070C0'), bottom=thin)
    cell_style(ws, r, 2, kh, align='center')
    cell_style(ws, r, 3, je, align='center')
    ws.row_dimensions[r].height = 22

out_path = r'C:\Users\Administrator\Desktop\智能体\ea策略\EA策略资金表.xlsx'
wb.save(out_path)
print(f"✅ 已保存: {out_path}")
