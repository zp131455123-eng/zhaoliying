# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "EA"

# 列宽
ws.column_dimensions['A'].width = 10
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 16
ws.column_dimensions['E'].width = 18

thin = Side(style='thin', color='000000')
medium = Side(style='medium', color='000000')
b_all = Border(left=thin, right=thin, top=thin, bottom=thin)
b_medium = Border(left=medium, right=medium, top=medium, bottom=medium)

def sc(row, col, value, bold=False, halign='center', valign='center', border=None, size=11, fill=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, size=size, name='Microsoft YaHei')
    c.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=True)
    if border:
        c.border = border
    if fill:
        c.fill = PatternFill(start_color=fill, end_color=fill, fill_type='solid')
    return c

# ===== 第1行：表头 =====
# A1: 负责人
sc(1, 1, '负责人', bold=True, border=b_all)
# B1-C1: 进行中（合并）
ws.merge_cells('B1:C1')
sc(1, 2, '进行中', bold=True, size=14, border=b_all)
ws.row_dimensions[1].height = 28

# ===== 进行中数据 row 2~18 =====
jinxing = [
    ('张鹏', '范宇轩', 2000),
    ('张鹏', '苏亮', 2000),
    ('张鹏', '雷熊', 2000),
    ('曾涛', '王欢', 2000),
    ('张鹏', '林锦豪', 4000),
    ('张鹏', '王泽远', 2000),
    ('张鹏', '梁杰煌', 2000),
    ('张鹏', '焦伟林', 10000),
    ('张鹏', '谢远然', 2000),
    ('公司', '雷蒙', 4000),
    ('张鹏', '聂平华', 6000),
    ('公司', '华哥（xiong yong）', 10000),
    ('刘涛', '孙成（chenyingying）', 2000),
    ('张鹏', '陈泽林', 2000),
    ('张鹏', '张金虎', 2000),
    ('刘涛', '杨静', 2000),
    ('张鹏', '杨浩', 2000),
]
jx_start = 2
jx_end = jx_start + len(jinxing) - 1  # 18

for i, (a, b, c) in enumerate(jinxing):
    r = jx_start + i
    sc(r, 1, a, border=b_all)
    sc(r, 2, b, border=b_all)
    sc(r, 3, c, border=b_all)
    ws.row_dimensions[r].height = 20

# D列合并：共计 58000（垂直居中，对应进行中中间行）
ws.merge_cells(start_row=jx_start, start_column=4, end_row=jx_end, end_column=4)
c = ws.cell(row=jx_start, column=4, value='共计：58000')
c.font = Font(bold=True, size=11, name='Microsoft YaHei')
c.alignment = Alignment(horizontal='center', vertical='center')
c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ===== 停止标题行 =====
stop_title = jx_end + 1  # 19
ws.merge_cells(start_row=stop_title, start_column=1, end_row=stop_title, end_column=3)
sc(stop_title, 1, '停止', bold=True, size=14, border=b_all)
ws.row_dimensions[stop_title].height = 28

# ===== 停止数据 =====
tingzhi = [
    ('刘涛', '雷伟平', 2000),
    ('刘涛', '朱睿', 2000),
    ('曾涛', '徐敏', 2000),
    ('张鹏', '孙明', 2000),
    ('张鹏', '李超杰', 2000),
    ('公司', '小叶', 8000),
    ('张鹏', '万诺', 2000),
    ('张鹏', '曾经，拥有', 2000),
    ('张鹏', '张凯嘉', 2000),
    ('张鹏', 'MN医疗', 2000),
    ('张鹏', 'bibiubiu', 2000),
    ('刘涛', '罗琴', 2000),
    ('张鹏', '徐宝龙', 2000),
]
tz_start = stop_title + 1  # 20
tz_end = tz_start + len(tingzhi) - 1  # 32

for i, (a, b, c) in enumerate(tingzhi):
    r = tz_start + i
    sc(r, 1, a, border=b_all)
    sc(r, 2, b, border=b_all)
    sc(r, 3, c, border=b_all)
    ws.row_dimensions[r].height = 20

# D列合并：共计 32000
ws.merge_cells(start_row=tz_start, start_column=4, end_row=tz_end, end_column=4)
c = ws.cell(row=tz_start, column=4, value='共计：32000')
c.font = Font(bold=True, size=11, name='Microsoft YaHei')
c.alignment = Alignment(horizontal='center', vertical='center')
c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

# E列合并：总计入金量 90000U（垂直居中，对应进行中+停止全部行）
ws.merge_cells(start_row=jx_start, start_column=5, end_row=tz_end, end_column=5)
c = ws.cell(row=jx_start, column=5, value='总计入金量：90000U')
c.font = Font(bold=True, size=12, name='Microsoft YaHei')
c.alignment = Alignment(horizontal='center', vertical='center')
c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ===== 空行 =====
empty_row = tz_end + 1  # 33
ws.row_dimensions[empty_row].height = 8

# ===== 贵2数据（带粗边框） =====
gui2 = [
    ('贵2', '李维', 2000),
    ('贵2', '刘学习', 10000),
    ('贵2', '刘学习推荐', 4000),
]
g2_start = empty_row + 1  # 34

for i, (a, b, c) in enumerate(gui2):
    r = g2_start + i
    # A列贵2用粗边框
    sc(r, 1, a, border=b_medium)
    sc(r, 2, b, border=b_all)
    sc(r, 3, c, border=b_all)
    ws.row_dimensions[r].height = 20

out = r'C:\Users\Administrator\Desktop\智能体\ea策略\EA策略资金表.xlsx'
wb.save(out)
