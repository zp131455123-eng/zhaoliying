#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
各平台卡户链接管理工具
用法:
  python links_manager.py list                        # 列出所有
  python links_manager.py list 星迈                   # 按平台筛选
  python links_manager.py add_client 平台 账户类型 姓名  # 添加开户客户
  python links_manager.py del_client 平台 账户类型 姓名  # 删除客户
  python links_manager.py export                      # 导出 Excel
"""
import json, sys, os

DATA_FILE = os.path.join(os.path.dirname(__file__), "links.json")

def load():
    if not os.path.exists(DATA_FILE):
        return []
    with open(DATA_FILE, encoding="utf-8") as f:
        return json.load(f)

def save(data):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def find(data, platform, account_type):
    for d in data:
        if d["platform"] == platform and d["account_type"] == account_type:
            return d
    return None

def cmd_list(platform=None):
    data = load()
    if platform:
        data = [d for d in data if platform in d["platform"]]
    if not data:
        print("没有数据")
        return
    platforms = {}
    for d in data:
        platforms.setdefault(d["platform"], []).append(d)
    for p, items in platforms.items():
        print(f"\n{'='*55}")
        print(f"  【{p}】")
        print(f"{'='*55}")
        for i in items:
            clients = i.get("clients", [])
            print(f"  [{i['account_type']}]  ({len(clients)}人开户)")
            print(f"  链接: {i['link']}")
            if i.get("note"):
                print(f"  备注: {i['note']}")
            if clients:
                print(f"  开户客户: {' / '.join(clients)}")
            print()

def cmd_add_client(platform, account_type, name):
    data = load()
    entry = find(data, platform, account_type)
    if not entry:
        print(f"未找到: {platform} - {account_type}")
        return
    if name in entry.get("clients", []):
        print(f"{name} 已在列表中")
        return
    entry.setdefault("clients", []).append(name)
    save(data)
    print(f"已添加: {platform} / {account_type} → {name}")

def cmd_del_client(platform, account_type, name):
    data = load()
    entry = find(data, platform, account_type)
    if not entry or name not in entry.get("clients", []):
        print(f"未找到: {name}")
        return
    entry["clients"].remove(name)
    save(data)
    print(f"已删除: {platform} / {account_type} → {name}")

def cmd_export():
    try:
        import openpyxl
    except ImportError:
        os.system("pip install openpyxl -q")
        import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    data = load()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "各平台卡户链接"

    # 标题行
    headers = ["平台", "账户类型", "账户链接", "备注", "开户客户", "资金(U)", "开户人数"]
    ws.append(headers)
    col_widths = [12, 16, 80, 25, 40, 12, 10]
    for i, w in enumerate(col_widths, 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = w
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1a3a5c")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 数据行
    row_idx = 2
    platforms = {}
    for d in data:
        platforms.setdefault(d["platform"], []).append(d)

    for p, items in platforms.items():
        start = row_idx
        for i in items:
            clients = i.get("clients", [])
            # 兼容旧格式（字符串）和新格式（对象）
            if clients and isinstance(clients[0], dict):
                names = "、".join(c["name"] for c in clients)
                total_amount = sum(c.get("amount", 0) for c in clients)
                count = len(clients)
            else:
                names = "、".join(clients)
                total_amount = ""
                count = len(clients)
            ws.append([
                i["platform"],
                i["account_type"],
                i["link"],
                i.get("note", ""),
                names,
                total_amount,
                count
            ])
            # 链接可点击
            cell = ws.cell(row=row_idx, column=3)
            if cell.value and cell.value.startswith("http"):
                cell.hyperlink = cell.value
                cell.font = Font(color="0563C1", underline="single")
            row_idx += 1
        # 平台列合并
        if row_idx - start > 1:
            ws.merge_cells(f"A{start}:A{row_idx-1}")
            ws[f"A{start}"].alignment = Alignment(horizontal="center", vertical="center")

    # 隔行底色
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if row[0].row % 2 == 0:
            for cell in row:
                if not cell.fill or cell.fill.fill_type == "none":
                    cell.fill = PatternFill("solid", fgColor="EEF4FB")

    out = os.path.join(os.path.dirname(__file__), "各平台卡户链接.xlsx")
    wb.save(out)
    print(f"已导出: {out}")

if __name__ == "__main__":
    args = sys.argv[1:]
    cmd = args[0] if args else "list"

    if cmd == "list":
        cmd_list(args[1] if len(args) > 1 else None)
    elif cmd == "add_client" and len(args) == 4:
        cmd_add_client(args[1], args[2], args[3])
    elif cmd == "del_client" and len(args) == 4:
        cmd_del_client(args[1], args[2], args[3])
    elif cmd == "export":
        cmd_export()
    else:
        print(__doc__)
